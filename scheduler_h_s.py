import pandas as pd
from ortools.sat.python import cp_model
from datetime import datetime, timedelta
import os
import math
import subprocess


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EMP_FILE = os.path.join(BASE_DIR, "Book(Employees)_01.xlsx")
HOLIDAY_FILE = os.path.join(BASE_DIR, "Holidaydata.xlsx")
EVENT_FILE = os.path.join(BASE_DIR, "EventsData.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "Final_Rota_MultiSheet.xlsx")
EVENT_SCRIPT = os.path.join(BASE_DIR, "eventapicall.py")

# Try to import openpyxl
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except:
    print("openpyxl unavailable")
# Try to import openpyxl for highlighting
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("[WARNING] 'openpyxl' not found. Highlighting disabled.")

# ==============================================================================
#                           BUSINESS CONTROL PANEL
# ==============================================================================

ENABLE_HYBRID_MODE = True
REVENUE_PER_STAFF = 800   

MIN_SHIFT_LENGTH = 6
MAX_SHIFT_LENGTH = 9
MIN_REST_HOURS = 12
MAX_CONSECUTIVE_DAYS = 6

ENABLE_STRICT_SECOND_START = True
SECOND_START_EARLIEST = 9   
SECOND_START_LATEST = 10    

ENABLE_RUSH_LOCK = True 
RUSH_TRIGGER = 2
MIN_RUSH_STAFF = 2

EVENING_START_HOUR = 17 
MORNING_END_HOUR = 12    

# ==============================================================================

def parse_fixed_shifts(raw):
    """
    Converts:
    Monday|09:00|17:00;Wednesday|12:00|20:00

    Into:
    {
        "Monday": (9,17),
        "Wednesday": (12,20)
    }
    """
    result = {}

    raw = str(raw).strip()

    if raw in ["", "nan", "None"]:
        return result

    for item in raw.split(";"):
        parts = item.split("|")

        if len(parts) != 3:
            continue

        day_name = parts[0].strip()

        try:
            start_h = int(parts[1].split(":")[0])
            end_h = int(parts[2].split(":")[0])

            if end_h == 0:
                end_h = 24

            result[day_name] = (start_h, end_h)

        except:
            pass

    return result

def run_event_tracker():
    print("--- 0. UPDATING EVENT INTELLIGENCE ---")

    try:
        if os.path.exists(EVENT_SCRIPT):

            result = subprocess.run(
                [sys.executable, EVENT_SCRIPT],
                capture_output=True,
                text=True,
                cwd=BASE_DIR
            )

            if result.returncode == 0:
                print("Event scan complete.")
            else:
                print("Event scan error:", result.stderr)

        else:
            print("eventapicall.py not found.")

    except Exception as e:
        print("Event scanner failed:", e)

    print("----------------------------------")

def apply_formatting_to_sheet(ws, daily_shop_hours, event_info_map, approved_holidays):
    """
    Applies colors and event headers to a specific worksheet.
    """
    # Define Fills
    fill_opener = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid") # Light Yellow
    fill_closer = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid") # Light Red
    fill_holiday = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # Light Green
    fill_event_header = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") # Gold
    
    # Map column index to date string based on this specific sheet's headers
    col_date_map = {}
    header_row = ws[1]
    
    for cell in header_row:
        val = str(cell.value)
        # Header format: "2025-11-24 (Mon)"
        if '(' in val:
            date_part = val.split(' ')[0]
            col_date_map[cell.column] = date_part
            
            # Highlight Header if Event
            if date_part in event_info_map:
                cell.fill = fill_event_header
                cell.value = f"{val} [Event: {event_info_map[date_part]}]"

    # Iterate Data Rows (starting row 2)
    for row in ws.iter_rows(min_row=2):
        # Get Employee ID from column A (Assuming it's 1st column)
        emp_id_cell = row[0] 
        emp_id = str(emp_id_cell.value)
        
        for cell in row:
            col_idx = cell.column
            if col_idx in col_date_map:
                date_str = col_date_map[col_idx]
                cell_val = str(cell.value)
                
                # 1. Holiday Check
                if (emp_id, date_str) in approved_holidays:
                    cell.fill = fill_holiday
                    continue 
                
                if cell_val == "OFF": continue

                # 2. Shift Check
                # Parse "06:00 - 14:00"
                if " - " in cell_val:
                    try:
                        t_start_str, t_end_str = cell_val.split(" - ")
                        s_h = int(t_start_str.split(":")[0])
                        e_h = int(t_end_str.split(":")[0])
                        if e_h == 0: e_h = 24
                        
                        # Get Shop Hours for this day
                        shop_open, shop_close = daily_shop_hours.get(date_str, (0, 0))
                        
                        # Priority: Closer (Red) > Opener (Yellow)
                        if e_h == shop_close:
                            cell.fill = fill_closer
                        elif s_h == shop_open:
                            cell.fill = fill_opener
                    except: pass

def solve_rota_final_v14(emp_file, holiday_file, target_weeks=None):
    # 1. RUN THE EVENT SCANNER
    run_event_tracker()

    print("--- 1. LOADING DATA ---")

    try:
        # Load Sheets
        df_emp = pd.read_excel(emp_file, sheet_name="Employees")
        df_shifts = pd.read_excel(emp_file, sheet_name="Shift Templates")
        df_sales = pd.read_excel(emp_file, sheet_name="Sales Forecast") 
        
        # --- CLEAN HEADERS ---
        df_shifts.columns = df_shifts.columns.str.strip()
        df_emp.columns = df_emp.columns.str.strip()
        df_sales.columns = df_sales.columns.str.strip()
        
        print(f"  > Employee Columns: {list(df_emp.columns)}")

    except Exception as e:
        print(f"CRITICAL ERROR: Could not read {emp_file}. Reason: {e}")
        return

    # Merge Sales
    df_shifts['Date'] = pd.to_datetime(df_shifts['Date'])
    df_sales['Date'] = pd.to_datetime(df_sales['Date'])
    df_shifts = pd.merge(df_shifts, df_sales[['Date', 'Total Sales']], on='Date', how='left')

    # --- FILTERING LOGIC ---
    if target_weeks:
        target_timestamps = pd.to_datetime(target_weeks)
        df_shifts['Week_Start_Date'] = df_shifts['Date'].apply(lambda x: x - timedelta(days=x.weekday()))
        df_shifts = df_shifts[df_shifts['Week_Start_Date'].isin(target_timestamps)].copy()
        df_shifts = df_shifts.drop(columns=['Week_Start_Date'])

    # Load Holidays
    approved_holidays = set() 
    if os.path.exists(holiday_file):
        try:
            df_hol = pd.read_excel(holiday_file, sheet_name='Data')
            df_hol.columns = df_hol.columns.str.strip()
            df_hol['Date'] = pd.to_datetime(df_hol['Date']).dt.normalize()
            df_hol['Employee ID'] = df_hol['Employee ID'].astype(str).str.strip()

            approved_df = df_hol[df_hol['Status'].astype(str).str.lower() == 'approved']
            for _, row in approved_df.iterrows():
                h_date = row['Date'].strftime('%Y-%m-%d')
                h_id = row['Employee ID']
                approved_holidays.add((h_id, h_date))

            print(f'  > Loaded {len(approved_holidays)} approved holidays.')
        except Exception as e:
            print(f'  > Warning: Holiday file error: {e}')

    # Load Events Info for Highlighting
    event_info_map = {}
    if os.path.exists(EVENT_FILE):
        try:
            df_ev = pd.read_excel(EVENT_FILE)
            for _, row in df_ev.iterrows():
                d_str = pd.to_datetime(row['Date']).strftime('%Y-%m-%d')
                t_str = str(row['Start Time'])
                event_info_map[d_str] = t_str
            print(f"  > Found {len(event_info_map)} high-impact event days to highlight.")
        except Exception as e:
            print(f"  > Warning reading EventsData.xlsx: {e}")

    df_emp['ID'] = df_emp['ID'].astype(str)
    employees = df_emp.to_dict('index')
    
    df_shifts['Week_Num'] = df_shifts['Date'].dt.isocalendar().week
    unique_weeks = df_shifts['Week_Num'].unique()
    
    # NEW: Store DataFrames per week instead of a flat list
    weekly_dataframes = {}

    # Store Shop Hours for Coloring later: { '2025-11-24': (7, 24) }
    daily_shop_hours = {}

    print(f"--- 2. SOLVING FOR {len(unique_weeks)} WEEKS ---")

    for week in unique_weeks:
        print(f"\nProcessing Week {week}...")
        week_data = df_shifts[df_shifts['Week_Num'] == week].copy()
        week_data = week_data.sort_values('Date')
        dates_in_order = week_data['Date'].dt.strftime('%Y-%m-%d').unique()
        
        # --- BUDGET CHECK ---
        weekly_budget_hours = 9999
        if 'Budget' in week_data.columns:
            val = week_data['Budget'].max()
            if not pd.isna(val) and val > 0:
                weekly_budget_hours = int(val)
                print(f"  > [CONSTRAINT] Weekly Budget: {weekly_budget_hours} hours")

        model = cp_model.CpModel()
        
        work = {}           
        start = {}          
        is_working_day = {} 
        daily_start_hour = {} 
        daily_end_hour = {}   
        
        emp_indices = df_emp.index.tolist()
        all_worked_hours_vars = []

        # 1. SETUP VARIABLES
        for idx in emp_indices:
            for i, row in week_data.iterrows():
                date_str = row['Date'].strftime('%Y-%m-%d')
                start_h = pd.to_datetime(str(row['Start'])).hour
                end_h = pd.to_datetime(str(row['End'])).hour
                if end_h == 0: end_h = 24
                
                # Store for coloring later
                daily_shop_hours[date_str] = (start_h, end_h)

                is_working_day[(idx, date_str)] = model.NewBoolVar(f'day_{idx}_{date_str}')
                daily_start_hour[(idx, date_str)] = model.NewIntVar(0, 24, f'start_h_{idx}_{date_str}')
                daily_end_hour[(idx, date_str)] = model.NewIntVar(0, 24, f'end_h_{idx}_{date_str}')
                
                for h in range(start_h, end_h):
                    work[(idx, date_str, h)] = model.NewBoolVar(f'w_{idx}_{date_str}_{h}')
                    start[(idx, date_str, h)] = model.NewBoolVar(f's_{idx}_{date_str}_{h}')
                    all_worked_hours_vars.append(work[(idx, date_str, h)])

        # 2. CONSTRAINTS (Budget, Holidays, Contracts, etc)
        # A. BUDGET
        model.Add(sum(all_worked_hours_vars) <= weekly_budget_hours)

        # B. HOLIDAYS
        # B2. FIXED WEEKLY SHIFTS
        for idx in emp_indices:

            emp = employees[idx]
            emp_id = str(emp['ID']).strip()

            enabled = str(emp.get('Fixed Shift Enabled', 'No')).strip()

            if enabled != "Yes":
                continue

            fixed_map = parse_fixed_shifts(
                emp.get('Fixed Weekly Shift', '')
            )

            if not fixed_map:
                continue

            for _, row in week_data.iterrows():

                date_obj = row['Date']
                date_str = date_obj.strftime('%Y-%m-%d')
                day_name = date_obj.day_name()

                # Holiday overrides fixed shift
                if (emp_id, date_str) in approved_holidays:
                    continue

                # Unavailable overrides fixed shift
                unavailable = str(
                    emp.get('Unavailable Days', '')
                )

                if unavailable != 'nan' and day_name in unavailable:
                    continue

                if day_name not in fixed_map:
                    continue

                fixed_start, fixed_end = fixed_map[day_name]

                shop_start = pd.to_datetime(
                    str(row['Start'])
                ).hour

                shop_end = pd.to_datetime(
                    str(row['End'])
                ).hour

                if shop_end == 0:
                    shop_end = 24

                # clamp inside shop hours
                fixed_start = max(fixed_start, shop_start)
                fixed_end = min(fixed_end, shop_end)

                if fixed_end <= fixed_start:
                    continue

                model.Add(
                    is_working_day[(idx, date_str)] == 1
                )

                for h in range(shop_start, shop_end):

                    if fixed_start <= h < fixed_end:
                        model.Add(
                            work[(idx, date_str, h)] == 1
                        )
                    else:
                        model.Add(
                            work[(idx, date_str, h)] == 0
                        )
        for idx in emp_indices:
            emp_id = employees[idx]['ID']
            for i, row in week_data.iterrows():
                date_str = row['Date'].strftime('%Y-%m-%d')
                if (emp_id, date_str) in approved_holidays:
                    model.Add(is_working_day[(idx, date_str)] == 0)
                    start_h = pd.to_datetime(str(row['Start'])).hour
                    end_h = pd.to_datetime(str(row['End'])).hour
                    if end_h == 0: end_h = 24
                    for h in range(start_h, end_h):
                        model.Add(work[(idx, date_str, h)] == 0)

        # C. WEEKLY CONTRACT (CAPACITY)
        for idx in emp_indices:
            emp = employees[idx]
            emp_id = emp['ID']
            
            available_days_count = 0
            for d in dates_in_order:
                unavailable_str = str(emp.get('Unavailable Days', ''))
                day_name = pd.to_datetime(d).day_name()
                is_holiday = (emp_id, d) in approved_holidays
                is_unavailable = (unavailable_str != 'nan' and day_name in unavailable_str)
                if not is_holiday and not is_unavailable:
                    available_days_count += 1
            
            max_physical_capacity = available_days_count * MAX_SHIFT_LENGTH
            
            try:
                if 'Minimum Contractual Hours' in emp:
                    original_min = int(emp['Minimum Contractual Hours'])
                elif 'Minimum Contractual Hours ' in emp:
                    original_min = int(emp['Minimum Contractual Hours '])
                else:
                    original_min = 0 
            except: original_min = 0

            adjusted_min = min(original_min, max_physical_capacity)
            original_max = int(emp.get('Max Weekly Hours', 40))

            total_hours_vars = [work[(idx, row['Date'].strftime('%Y-%m-%d'), h)] 
                                for _, row in week_data.iterrows() 
                                for h in range(pd.to_datetime(str(row['Start'])).hour, 
                                               24 if pd.to_datetime(str(row['End'])).hour == 0 else pd.to_datetime(str(row['End'])).hour)]
            
            model.Add(sum(total_hours_vars) >= adjusted_min)
            model.Add(sum(total_hours_vars) <= original_max)
            
            total_working_days = [is_working_day[(idx, d)] for d in dates_in_order]
            model.Add(sum(total_working_days) <= MAX_CONSECUTIVE_DAYS) 

        # D. DAILY LOGIC
        for i, row in week_data.iterrows():
            date_str = row['Date'].strftime('%Y-%m-%d')
            start_h = pd.to_datetime(str(row['Start'])).hour
            end_h = pd.to_datetime(str(row['End'])).hour
            if end_h == 0: end_h = 24
            hours_range = list(range(start_h, end_h))

            for idx in emp_indices:
                model.Add(sum(work[(idx, date_str, h)] for h in hours_range) > 0).OnlyEnforceIf(is_working_day[(idx, date_str)])
                model.Add(sum(work[(idx, date_str, h)] for h in hours_range) == 0).OnlyEnforceIf(is_working_day[(idx, date_str)].Not())
                for h in hours_range:
                    if h == start_h: model.Add(start[(idx, date_str, h)] == work[(idx, date_str, h)])
                    else:
                        model.Add(work[(idx, date_str, h)] >= start[(idx, date_str, h)])
                        model.Add(start[(idx, date_str, h)] >= work[(idx, date_str, h)] - work[(idx, date_str, h-1)])
                model.Add(sum(start[(idx, date_str, h)] for h in hours_range) <= 1)
                model.Add(daily_start_hour[(idx, date_str)] == sum(h * start[(idx, date_str, h)] for h in hours_range))
                duration = sum(work[(idx, date_str, h)] for h in hours_range)
                model.Add(daily_end_hour[(idx, date_str)] == daily_start_hour[(idx, date_str)] + duration)
                model.Add(duration >= MIN_SHIFT_LENGTH).OnlyEnforceIf(is_working_day[(idx, date_str)])
                model.Add(duration <= MAX_SHIFT_LENGTH).OnlyEnforceIf(is_working_day[(idx, date_str)])

        # Fixed Slots
        for idx in emp_indices:
            emp = employees[idx]
            fixed_slot = str(emp.get('Fixed Slot', ''))
            if fixed_slot not in ['nan', 'None', 'Any']:
                for i, row in week_data.iterrows():
                    date_str = row['Date'].strftime('%Y-%m-%d')
                    start_h = pd.to_datetime(str(row['Start'])).hour
                    end_h = pd.to_datetime(str(row['End'])).hour
                    if end_h == 0: end_h = 24
                    if 'Morning' in fixed_slot and 'Evening' not in fixed_slot:
                        for h in range(start_h, end_h):
                            if h >= EVENING_START_HOUR: model.Add(work[(idx, date_str, h)] == 0)
                    if 'Evening' in fixed_slot and 'Morning' not in fixed_slot:
                        for h in range(start_h, end_h):
                            if h < MORNING_END_HOUR: model.Add(work[(idx, date_str, h)] == 0)
                    if fixed_slot == 'Afternoon':
                        for h in range(start_h, end_h):
                            if h < 10 or h >= 20: model.Add(work[(idx, date_str, h)] == 0)

        # Rest Period
        for idx in emp_indices:
            for i in range(len(dates_in_order) - 1):
                today = dates_in_order[i]
                tomorrow = dates_in_order[i+1]
                model.Add((daily_start_hour[(idx, tomorrow)] + 24) - daily_end_hour[(idx, today)] >= MIN_REST_HOURS).OnlyEnforceIf([is_working_day[(idx, today)], is_working_day[(idx, tomorrow)]])

        # E. STAFFING LOGIC (Hybrid)
        for i, row in week_data.iterrows():
            date_str = row['Date'].strftime('%Y-%m-%d')
            day_name = row['Date'].day_name()
            manual_min = int(row['Minimum Staff'])
            
            sales_min = 0
            if ENABLE_HYBRID_MODE:
                sales_val = row['Total Sales']
                if pd.isna(sales_val): sales_val = 0
                sales_min = math.ceil(sales_val / REVENUE_PER_STAFF)
            
            final_min_headcount = max(manual_min, sales_min)
            min_closing = int(row['Minimum closing staff'])
            min_headcount = max(final_min_headcount, min_closing)
            
            manual_max = int(row['Maximum Employees'])
            max_headcount = max(manual_max, min_headcount) 

            # Apply
            model.Add(sum(is_working_day[(idx, date_str)] for idx in emp_indices) >= min_headcount)
            model.Add(sum(is_working_day[(idx, date_str)] for idx in emp_indices) <= max_headcount)
            
            start_h = pd.to_datetime(str(row['Start'])).hour
            end_h = pd.to_datetime(str(row['End'])).hour
            if end_h == 0: end_h = 24
            for h in range(start_h, end_h):
                model.Add(sum(work[(idx, date_str, h)] for idx in emp_indices) >= 1)
            last_hour = end_h - 1
            model.Add(sum(work[(idx, date_str, last_hour)] for idx in emp_indices) >= min_closing)

            # --- UPDATED STRICT SECOND START LOGIC ---
            if ENABLE_STRICT_SECOND_START:
                if day_name == 'Sunday' and start_h == 6 and max_headcount >= 2:
                    model.Add(sum(work[(idx, date_str, 6)] for idx in emp_indices) == 1)
                    model.Add(sum(work[(idx, date_str, 7)] for idx in emp_indices) >= 2)
                elif start_h < SECOND_START_EARLIEST and end_h > SECOND_START_LATEST and max_headcount >= 2:
                    for h in range(start_h, SECOND_START_EARLIEST):
                         model.Add(sum(work[(idx, date_str, h)] for idx in emp_indices) == 1)
                    model.Add(sum(work[(idx, date_str, SECOND_START_LATEST)] for idx in emp_indices) >= 2)

            if ENABLE_RUSH_LOCK:
                operating_hours = list(range(start_h, end_h))
                for k in range(1, len(operating_hours) - 1):
                    current_h = operating_hours[k]
                    prev_h = operating_hours[k-1]
                    current_staff = sum(work[(idx, date_str, current_h)] for idx in emp_indices)
                    prev_staff = sum(work[(idx, date_str, prev_h)] for idx in emp_indices)
                    rush_active = model.NewBoolVar(f'rush_active_{date_str}_{prev_h}')
                    model.Add(prev_staff >= RUSH_TRIGGER).OnlyEnforceIf(rush_active)
                    model.Add(prev_staff < RUSH_TRIGGER).OnlyEnforceIf(rush_active.Not())
                    model.Add(current_staff >= MIN_RUSH_STAFF).OnlyEnforceIf(rush_active)

        # G. OPENER
        for i, row in week_data.iterrows():
            date_str = row['Date'].strftime('%Y-%m-%d')
            start_h = pd.to_datetime(str(row['Start'])).hour
            trained = [idx for idx in emp_indices if employees[idx].get('Opening Trained', 'No') == 'Yes']
            model.Add(sum(work[(idx, date_str, start_h)] for idx in trained) >= 1)
            
        for idx in emp_indices:
            emp = employees[idx]
            unavailable_str = str(emp.get('Unavailable Days', ''))
            for i, row in week_data.iterrows():
                date_str = row['Date'].strftime('%Y-%m-%d')
                if unavailable_str != 'nan' and row['Date'].day_name() in unavailable_str:
                    model.Add(is_working_day[(idx, date_str)] == 0)

        # 3. OBJECTIVE
        objective_terms = []
        for idx in emp_indices:
            pref = str(employees[idx].get('Preferred Day', ''))
            for i, row in week_data.iterrows():
                date_str = row['Date'].strftime('%Y-%m-%d')
                if pref != 'nan' and row['Date'].day_name() in pref:
                     objective_terms.append(is_working_day[(idx, date_str)])
        model.Maximize(sum(objective_terms))

        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 45.0
        status = solver.Solve(model)
        
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            print(f"  > Schedule found for Week {week}")
            
            # --- EXTRACT DATA FOR THIS WEEK ---
            current_week_results = []
            
            for idx in emp_indices:
                row_data = {
                    'Employee ID': employees[idx]['ID'],
                    'Name': employees[idx]['Name'],
                    # We don't need 'Week' in row data anymore as the Sheet Name indicates it
                }
                weekly_h = 0
                for i, row in week_data.iterrows():
                    date_str = row['Date'].strftime('%Y-%m-%d')
                    start_h = pd.to_datetime(str(row['Start'])).hour
                    end_h = pd.to_datetime(str(row['End'])).hour
                    if end_h == 0: end_h = 24
                    
                    shift_start, shift_end = None, None
                    for h in range(start_h, end_h):
                        if solver.Value(work[(idx, date_str, h)]):
                            if shift_start is None: shift_start = h
                            shift_end = h + 1
                            weekly_h += 1
                    col_key = f"{date_str} ({row['Date'].day_name()[:3]})"
                    if shift_start is not None:
                        row_data[col_key] = f"{shift_start:02d}:00 - {shift_end:02d}:00"
                    else:
                        emp_id = str(employees[idx]['ID']).strip()

                        if (emp_id, date_str) in approved_holidays:
                            row_data[col_key] = "Holiday"
                        else:
                            row_data[col_key] = "OFF"
                row_data['Total Weekly Hours'] = weekly_h
                current_week_results.append(row_data)
            
            # Create DataFrame for this week and store it in dictionary
            df_week = pd.DataFrame(current_week_results)
            # Reorder columns to put Name/Total first
            cols = ['Name', 'Total Weekly Hours'] + [c for c in df_week.columns if c not in ['Name', 'Employee ID', 'Total Weekly Hours']]
            df_week = df_week[cols]
            
            weekly_dataframes[week] = df_week
            
        else:
            print(f"  > NO SOLUTION for Week {week}. Budget too tight?")

    # --- WRITE TO EXCEL (MULTI-SHEET) ---
    if weekly_dataframes:
        output_file = OUTPUT_FILE
        print(f"\n--- WRITING TO EXCEL: {output_file} ---")
        
        # 1. Load ALL existing sheets from the file (if it exists), so we don't lose them
        existing_sheets = {}
        if os.path.exists(output_file):
            try:
                existing_wb = load_workbook(output_file)
                for sname in existing_wb.sheetnames:
                    ws = existing_wb[sname]
                    data = ws.values
                    headers = next(data, None)
                    if headers:
                        existing_sheets[sname] = pd.DataFrame(data, columns=headers)
                    else:
                        existing_sheets[sname] = pd.DataFrame()
                print(f"  > Loaded {len(existing_sheets)} existing sheet(s): {list(existing_sheets.keys())}")
            except Exception as e:
                print(f"  > Warning: Could not load existing file ({e}). Starting fresh.")
 
        # 2. Merge: new weeks overwrite their sheet, all other existing sheets are kept
        new_sheet_names = set()
        for week_num, df_week in weekly_dataframes.items():
            sheet_name = f"Week {week_num}"
            existing_sheets[sheet_name] = df_week   # add or replace
            new_sheet_names.add(sheet_name)
            print(f"  > Prepared {sheet_name} (new/updated)")
 
        # 3. Write all sheets (existing + new) back to file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df_sheet in existing_sheets.items():
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
 
        # 4. Apply formatting only to newly generated sheets
        try:
            print("  > Applying Highlights to new sheet(s)...")
            wb = load_workbook(output_file)
            for sheet_name in new_sheet_names:
                if sheet_name in wb.sheetnames:
                    print(f"    - Formatting {sheet_name}...")
                    ws = wb[sheet_name]
                    apply_formatting_to_sheet(ws, daily_shop_hours, event_info_map, approved_holidays)
            wb.save(output_file)
            print(f"  > Formatting complete.")
        except Exception as e:
            print(f"  > Formatting Error: {e}")
 
        print(f"\nSUCCESS! Rota saved. Total sheets in file: {len(existing_sheets)}")
    else:
        print("No Rota generated for any week.")

# Run
if __name__ == "__main__":
    solve_rota_final_v14(
        EMP_FILE,
        HOLIDAY_FILE
    )
